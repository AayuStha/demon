import requests
from datetime import datetime, timedelta
from dotenv import load_dotenv
import pytz
import re
import os
load_dotenv()
API_KEY = os.getenv("API_KEY")

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# Your GitHub Personal Access Token (replace with your actual token)
GITHUB_ACCESS_TOKEN = "API_KEY"

# Function to retrieve first and last commit dates
def get_first_and_last_commit(repo_owner, repo_name):
    url = f'https://api.github.com/repos/{repo_owner}/{repo_name}/commits'
    headers = {
        "Authorization": f"token {GITHUB_ACCESS_TOKEN}"
    }
    params = {'per_page': 100, 'page': 1}
    commits = []

    while True:
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            new_commits = response.json()
            if not new_commits:
                break
            commits.extend(new_commits)
            params['page'] += 1
        elif response.status_code == 403:  # Rate limit exceeded or unauthorized
            return None, None, "Rate limit exceeded or unauthorized"
        else:
            return None, None, f"Error: {response.status_code}"

    if not commits:
        return None, None, "No commits found"

    first_commit = commits[-1]['commit']['author']['date']
    last_commit = commits[0]['commit']['author']['date']
    return len(commits), first_commit, last_commit

# Convert UTC timestamp to Kathmandu timezone
def convert_utc_to_kathmandu(utc_timestamp):
    utc_datetime = datetime.strptime(utc_timestamp, '%Y-%m-%dT%H:%M:%SZ')
    utc_datetime = utc_datetime.replace(tzinfo=pytz.utc)
    kathmandu_timezone = pytz.timezone('Asia/Kathmandu')
    return utc_datetime.astimezone(kathmandu_timezone)

# Extract username and repository name from GitHub URL
def extract_repo_user(url):
    pattern = r'https://github.com/([^/]+)/([^/?]+?)(?:\.git)?$'
    match = re.match(pattern, url)
    if match:
        return match.group(1), match.group(2)
    return None, None

# Define hackathon times
kathmandu_timezone = pytz.timezone('Asia/Kathmandu')
hackathon_start = kathmandu_timezone.localize(datetime(2024, 11, 28, 8, 30, 0))
hackathon_end = hackathon_start + timedelta(hours=48)

# Example GitHub URLs
github_urls = [
"https://github.com/tiixsha/byte_me",
"https://github.com/shristee-mah/AI-Valanche",
"https://github.com/SicaDeveloper/Nursesnearme",
"https://github.com/Hunter-420/NHA-National-Health-Agency",
"https://github.com/darshan2436/dirghaayu",
"https://github.com/mahatoankit/NeuroVision",
"https://github.com/prathama7/Orbit-Hackathon",
"https://github.com/upendrapant/civishub.git",
"https://github.com/TinyCoders-q-q/HealConnect",
"https://github.com/stha-sanket/LENS_OCR-digitalizer_ObrbitHack.git",
"https://github.com/pranjalaryal22/The-real-one",
"https://github.com/Nirakar1414/Obrit_hacks",
"https://github.com/GhimirePlan/Team-Satya-AI-Powered-Fake-News-Detection-Tool-for-Nepali-Media",
"https://github.com/zzenn44/Vaxtrack",
"https://github.com/Puspa222/sanskala",
"https://github.com/SudipTimalsina/Orbit-hacks",
"https://github.com/anujpaude1/dementia",
"https://github.com/bikashmishraa/bikash-ResQCode.git",
"https://github.com/saugatBhupal/Krishi-Sathi.git",
"https://github.com/GrishmaDon/BMI-Calculator.git",
"https://github.com/Kaveri17/Bio-Medical-Pet-Care-",
"https://github.com/Darshan808/Sanket-App",
"https://github.com/MadanBelbase/Orbit_2024_Hackathon",
"https://github.com/legendofzer0/orbit-hack-Team-7",
"https://github.com/rasadregmi/ClassLens",
"https://github.com/kritan-pixel/Code-Odyssey.git",
"https://github.com/prabhakarKandel69/Hackathon-BitEr.git",
"https://github.com/Paracetamol77/Hackathon-1.git",
"https://github.com/Prasun-Shiwakoti/CoCo",
"https://github.com/adhikari-arpan/OrbitHacks2024_Iterators"
]

outputs = []
try:
    for github_url in github_urls:
        username, repo = extract_repo_user(github_url)
        if not username or not repo:
            outputs.append({
                "repo": None,
                "username": None,
                "status": "Invalid URL"
            })
            continue

        try:
            total_commits, first_commit, last_commit = get_first_and_last_commit(username, repo)
            if not total_commits:
                outputs.append({
                    "repo": repo,
                    "username": username,
                    "status": first_commit  # Error message
                })
                continue

            first_commit_kathmandu = convert_utc_to_kathmandu(first_commit)
            last_commit_kathmandu = convert_utc_to_kathmandu(last_commit)

            # Check if commits exist before hackathon
            has_committed_before_hackathon = first_commit_kathmandu < hackathon_start

            # Count commits during hackathon
            commits_during_hackathon = 0
            if first_commit_kathmandu <= hackathon_end:
                commits_during_hackathon = sum(
                    hackathon_start <= convert_utc_to_kathmandu(commit['commit']['author']['date']) <= hackathon_end
                    for commit in requests.get(
                        f"https://api.github.com/repos/{username}/{repo}/commits",
                        headers={"Authorization": f"token {GITHUB_ACCESS_TOKEN}"},
                        params={"per_page": 100, "page": 1}
                    ).json()
                )

            # Append output
            outputs.append({
                "repo": repo,
                "username": username,
                "commits": total_commits,
                "first_commit_kathmandu": first_commit_kathmandu.isoformat(),
                "last_commit_kathmandu": last_commit_kathmandu.isoformat(),
                "has_committed_before_hackathon": has_committed_before_hackathon,
                "commits_during_hackathon": commits_during_hackathon,
                "hackathon_start": hackathon_start.isoformat(),
                "hackathon_end": hackathon_end.isoformat(),
                "status": 200
            })
        except Exception as e:
            outputs.append({
                "repo": repo,
                "username": username,
                "status": f"Error: {str(e)}"
            })
finally:
    # Write to Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hackathon Analysis"

    # Add timestamp to the top of the sheet
    current_time_kathmandu = datetime.now(kathmandu_timezone).strftime("%Y-%m-%d %H:%M:%S")
    sheet["A1"] = f"Data Generated On: {current_time_kathmandu}"
    sheet.merge_cells("A1:H1")  # Merge cells for better formatting
    sheet["A1"].style = "Title"  # Apply a title style for visibility

    # Add headers
    headers = [
        "Repo", "Username", "Commits", "First Commit", "Last Commit",
        "Commits During Hackathon", "Before Hackathon?", "Status"
    ]
    sheet.append(headers)

    # Add data to the sheet
    for output in outputs:
        row = [
            output.get("repo"),
            output.get("username"),
            output.get("commits"),
            output.get("first_commit_kathmandu"),
            output.get("last_commit_kathmandu"),
            output.get("commits_during_hackathon"),
            output.get("has_committed_before_hackathon"),
            output.get("status")
        ]
        sheet.append(row)

    # Apply gradient coloring for "Commits" column (Column C)
    max_row = sheet.max_row
    color_rule = ColorScaleRule(
        start_type="min", start_color="FFCCCC",
        mid_type="percentile", mid_value=50, mid_color="FFFF99",
        end_type="max", end_color="99FF99"
    )
    sheet.conditional_formatting.add(f"C2:C{max_row}", color_rule)

    # Save the Excel file
    workbook.save("Hackathon_Analysis_Colored.xlsx")
