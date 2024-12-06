import requests
from datetime import datetime, timedelta
from dotenv import load_dotenv
import pytz
import re
import os
load_dotenv()
API_KEY = os.getenv("API_KEY")

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# Your GitHub Personal Access Token (replace with your actual token)
GITHUB_ACCESS_TOKEN = "API_KEY"

# Mapping for team names and repository links
team_data = {
    "https://github.com/tiixsha/byte_me": "byte me",
    "https://github.com/shristee-mah/AI-Valanche": "AI-valanche",
    "https://github.com/SicaDeveloper/Nursesnearme": "Script Kiddies",
    "https://github.com/Hunter-420/NHA-National-Health-Agency": "cosmIT",
    "https://github.com/darshan2436/dirghaayu": "ERROR 404",
    "https://github.com/mahatoankit/NeuroVision": "NeuroVision",
    "https://github.com/prathama7/Orbit-Hackathon": "Bug Busters",
    "https://github.com/upendrapant/civishub.git": "Sambhala",
    "https://github.com/TinyCoders-q-q/HealConnect": "Tiny Coders",
    "https://github.com/stha-sanket/LENS_OCR-digitalizer_ObrbitHack.git": "Lens",
    "https://github.com/pranjalaryal22/The-real-one": "The Real One",
    "https://github.com/Nirakar1414/Obrit_hacks": "Cache me if you can",
    "https://github.com/GhimirePlan/Team-Satya-AI-Powered-Fake-News-Detection-Tool-for-Nepali-Media": "Team Satya",
    "https://github.com/zzenn44/Vaxtrack": "De-v's",
    "https://github.com/Puspa222/sanskala": "Hamro Sanskala",
    "https://github.com/SudipTimalsina/Orbit-hacks": "Tech Sparks LEC",
    "https://github.com/anujpaude1/dementia": "TypeError",
    "https://github.com/bikashmishraa/bikash-ResQCode.git": "ResQCode",
    "https://github.com/saugatBhupal/Krishi-Sathi.git": "Team Mato",
    "https://github.com/GrishmaDon/BMI-Calculator.git": "Infinity Squad",
    "https://github.com/Kaveri17/Bio-Medical-Pet-Care-": "code bridge",
    "https://github.com/Darshan808/Sanket-App": "Team Sanket",
    "https://github.com/MadanBelbase/Orbit_2024_Hackathon": "checkMate",
    "https://github.com/legendofzer0/orbit-hack-Team-7": "Team-7",
    "https://github.com/rasadregmi/ClassLens": "Signify",
    "https://github.com/kritan-pixel/Code-Odyssey.git": "Code Odyssey",
    "https://github.com/prabhakarKandel69/Hackathon-BitEr.git": "Bit-Er",
    "https://github.com/Paracetamol77/Hackathon-1.git": "Omkara",
    "https://github.com/Prasun-Shiwakoti/CoCo": "Sentinels",
    "https://github.com/adhikari-arpan/OrbitHacks2024_Iterators": "Team Iterators",
}

# Function to retrieve commits across all branches
def get_all_branch_commits(repo_owner, repo_name):
    url = f'https://api.github.com/repos/{repo_owner}/{repo_name}/commits'
    headers = {"Authorization": f"token {GITHUB_ACCESS_TOKEN}"}
    params = {'per_page': 100, 'page': 1}
    commit_dates = []

    while True:
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            commits = response.json()
            if not commits:
                break
            commit_dates.extend([commit['commit']['author']['date'] for commit in commits])
            params['page'] += 1
        elif response.status_code == 403:  # Rate limit exceeded or unauthorized
            return None, "Rate limit exceeded or unauthorized"
        else:
            return None, f"Error: {response.status_code}"

    return len(commit_dates), commit_dates

# Convert UTC timestamp to Kathmandu timezone
def convert_utc_to_kathmandu(utc_timestamp):
    utc_datetime = datetime.strptime(utc_timestamp, '%Y-%m-%dT%H:%M:%SZ')
    utc_datetime = utc_datetime.replace(tzinfo=pytz.utc)
    kathmandu_timezone = pytz.timezone('Asia/Kathmandu')
    return utc_datetime.astimezone(kathmandu_timezone)

# Calculate commit rate
def calculate_commit_rate(commit_dates, hackathon_start, hackathon_end):
    kathmandu_commit_dates = [
        convert_utc_to_kathmandu(date) for date in commit_dates
    ]
    hackathon_commits = [
        date for date in kathmandu_commit_dates
        if hackathon_start <= date <= hackathon_end
    ]
    hackathon_duration = (hackathon_end - hackathon_start).total_seconds() / 3600  # hours
    commit_rate = len(hackathon_commits) / hackathon_duration
    return len(hackathon_commits), commit_rate

# Extract username and repository name from GitHub URL
def extract_repo_user(url):
    pattern = r'https://github.com/([^/]+)/([^/?]+?)(?:\.git)?$'
    match = re.match(pattern, url)
    if match:
        return match.group(1), match.group(2)
    return None, None

# Define hackathon times
kathmandu_timezone = pytz.timezone('Asia/Kathmandu')
hackathon_start = kathmandu_timezone.localize(datetime(2024, 11, 28, 8, 30, 0))
hackathon_end = hackathon_start + timedelta(hours=48)

# Example GitHub URLs
github_urls = list(team_data.keys())

outputs = []
try:
    for github_url in github_urls:
        username, repo = extract_repo_user(github_url)
        if not username or not repo:
            outputs.append({
                "repo": None,
                "username": None,
                "team_name": None,
                "status": "Invalid URL"
            })
            continue

        team_name = team_data.get(github_url, "Unknown Team")  # Get team name or mark as unknown

        try:
            total_commits, commit_dates = get_all_branch_commits(username, repo)
            if total_commits is None:
                outputs.append({
                    "repo": repo,
                    "username": username,
                    "team_name": team_name,
                    "status": commit_dates  # Error message
                })
                continue
            
            first_commit_kathmandu = convert_utc_to_kathmandu(min(commit_dates))
            last_commit_kathmandu = convert_utc_to_kathmandu(max(commit_dates))

            # Check commits before hackathon
            has_committed_before_hackathon = first_commit_kathmandu < hackathon_start

            # Commits and rate during hackathon
            commits_during_hackathon, commit_rate = calculate_commit_rate(
                commit_dates, hackathon_start, hackathon_end
            )

            # Append output
            outputs.append({
                "repo": repo,
                "username": username,
                "team_name": team_name,
                "commits": total_commits,
                "first_commit_kathmandu": first_commit_kathmandu.isoformat(),
                "last_commit_kathmandu": last_commit_kathmandu.isoformat(),
                "has_committed_before_hackathon": has_committed_before_hackathon,
                "commits_during_hackathon": commits_during_hackathon,
                "commit_rate_per_hour": round(commit_rate, 2),
                "hackathon_start": hackathon_start.isoformat(),
                "hackathon_end": hackathon_end.isoformat(),
                "status": 200
            })
        except Exception as e:
            outputs.append({
                "repo": repo,
                "username": username,
                "team_name": team_name,
                "status": f"Error: {str(e)}"
            })
finally:
    # Write to Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hackathon Analysis"

    # Add timestamp to the top of the sheet
    current_time_kathmandu = datetime.now(kathmandu_timezone).strftime("%Y-%m-%d %H:%M:%S")
    sheet["A1"] = f"Data Generated On: {current_time_kathmandu}"
    sheet.merge_cells("A1:J1")  # Adjust for additional column
    sheet["A1"].style = "Title"  # Apply a title style for visibility

    # Add headers
    headers = [
        "Team Name", "Repo", "Username", "Commits", "First Commit", "Last Commit",
        "Commits During Hackathon", "Commit Rate (Per Hour)", "Before Hackathon?", "Status"
    ]
    sheet.append(headers)

    # Add data
    for output in outputs:
        sheet.append([
            output.get("team_name", "Unknown"),
            output.get("repo", "N/A"),
            output.get("username", "N/A"),
            output.get("commits", 0),
            output.get("first_commit_kathmandu", "N/A"),
            output.get("last_commit_kathmandu", "N/A"),
            output.get("commits_during_hackathon", 0),
            output.get("commit_rate_per_hour", 0.0),
            output.get("has_committed_before_hackathon", "N/A"),
            output.get("status", "N/A")
        ])

    # Get the maximum row for formatting
    max_row = sheet.max_row

     # Apply conditional formatting to "Commits" (Column D)
    commits_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for row in range(2, max_row + 1):
        cell_value = sheet[f"D{row}"].value
        if cell_value and isinstance(cell_value, int) and cell_value > 0:
            sheet[f"D{row}"].fill = commits_fill

    # Apply conditional formatting to "Commits During Hackathon" (Column G)
    hackathon_commit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for row in range(2, max_row + 1):
        cell_value = sheet[f"G{row}"].value
        if cell_value and isinstance(cell_value, int) and cell_value > 0:
            sheet[f"G{row}"].fill = hackathon_commit_fill

    # Apply conditional formatting for errors in "Status" (Column J)
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in range(2, max_row + 1):
        cell_value = sheet[f"J{row}"].value
        if cell_value and "Error" in str(cell_value):
            sheet[f"J{row}"].fill = error_fill

# Apply conditional formatting to "Commit Rate (Per Hour)" column (Column H)
max_row = sheet.max_row
sheet.conditional_formatting.add(
    f"H2:H{max_row}",
    ColorScaleRule(start_type="min", start_color="FFAAAA", end_type="max", end_color="AAFFAA")
)

# Apply fill colors based on "True" or "False" in column I (Before Hackathon?)
for row in range(2, max_row + 1):
    cell = sheet[f"I{row}"]
    if str(cell.value).upper() == "TRUE":  # Compare with uppercase string
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif str(cell.value).upper() == "FALSE":  # Compare with uppercase string
        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Save the workbook to a file
    workbook.save("Hackathon_Analysis.xlsx")
    print("Data written to 'Hackathon_Analysis.xlsx'")
